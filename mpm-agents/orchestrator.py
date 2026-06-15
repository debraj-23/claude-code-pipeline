"""
orchestrator.py
===============

Fully autonomous Orchestrator agent for the MPM migration pipeline.

Author: Debraj Som

What it does
------------
1. Reads requirements from an Excel workbook (openpyxl).
2. Decomposes them and assigns each to the right subagent.
3. Calls the three subagents - figma_agent, spec_agent, test_agent - to do the work.
4. Lets you monitor progress at the subagent level (live status table + agent_status.json).

Two dispatch modes
------------------
* LLM mode (default): the orchestrator is itself a Claude agent. The three
  subagents are exposed to it as TOOLS, so Claude reads the requirements,
  decides decomposition/ordering, and calls the subagents autonomously. Needs
  ANTHROPIC_API_KEY.
* Deterministic mode (--no-llm): the orchestrator routes each requirement to a
  subagent by its `Type`/`Target` columns and calls them directly - no API key
  needed. Used by validate.py (with MPM_AGENTS_MOCK=1) to prove the wiring.

Both modes call the SAME run_figma_agent / run_spec_agent / run_test_agent
functions, so "the subagents are called from the orchestrator" holds either way.

Usage
-----
    python orchestrator.py --requirements requirements.xlsx
    python orchestrator.py --requirements requirements.xlsx --no-llm
    python orchestrator.py --init-sample        # write a sample requirements.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import List, Optional

_FIGMA_URL_RE = re.compile(r"https?://(?:www\.)?figma\.com/\S+")


def _extract_figma_url(*texts: str) -> Optional[str]:
    """Pull the first Figma URL out of any of the given strings (e.g. Notes)."""
    for t in texts:
        if not t:
            continue
        m = _FIGMA_URL_RE.search(t)
        if m:
            return m.group(0).rstrip(".,;)")
    return None

from sdk_runner import (
    AgentRunner,
    StatusTracker,
    Tool,
    McpStdioBridge,
    load_figma_mcp_bridge,
    MPM_UI_DIR,
    MPM_BACKEND_DIR,
    UI_REL,
    BACKEND_REL,
    AGENTS_DIR,
    DEFAULT_SPEC_FILE,
    DEFAULT_STANDARDS_FILE,
    REVIEWS_REL,
    MOCK,
    print_banner,
)
import figma_agent
import spec_agent
import test_agent
import review_agent

ORCH_NAME = "orchestrator"
DEFAULT_REQUIREMENTS = AGENTS_DIR / "requirements.xlsx"

# Columns expected in the requirements workbook (header row 1).
COLUMNS = ["ID", "Requirement", "Type", "Target", "Priority", "Notes"]


# ---------------------------------------------------------------------------
# Requirements I/O
# ---------------------------------------------------------------------------

def read_requirements(path: Path) -> List[dict]:
    """Read the requirements workbook into a list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    reqs: List[dict] = []
    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        record = {}
        for i, col in enumerate(header):
            record[col] = raw[i] if i < len(raw) and raw[i] is not None else ""
        # Normalise the known fields with case-insensitive header matching.
        norm = {k.lower(): v for k, v in record.items()}
        reqs.append({
            "id": str(norm.get("id", "")).strip(),
            "requirement": str(norm.get("requirement", "")).strip(),
            "type": str(norm.get("type", "auto")).strip().lower() or "auto",
            "target": str(norm.get("target", "")).strip().lower(),
            "priority": str(norm.get("priority", "")).strip(),
            "notes": str(norm.get("notes", "")).strip(),
        })
    return reqs


def requirements_to_markdown(reqs: List[dict]) -> str:
    lines = ["| ID | Requirement | Type | Target | Priority | Notes |",
             "|----|-------------|------|--------|----------|-------|"]
    for r in reqs:
        lines.append(
            f"| {r['id']} | {r['requirement']} | {r['type']} | {r['target']} | {r['priority']} | {r['notes']} |"
        )
    return "\n".join(lines)


def write_sample_requirements(path: Path) -> None:
    """Create a sample requirements.xlsx so the pipeline can be run end-to-end."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append(COLUMNS)
    fig = "https://www.figma.com/design/e2RQMTHDbU2eSIKrTFXoAf/Untitled?node-id="
    sample = [
        ["R1", "Generate the Spring Boot backend (auth, organisation CRUD, JWT, roles) from specs.md",
         "spec", "backend", "High", "Use grails-ui-demo/specs.md as the spec."],
        ["R2", "Convert the Login screen Figma design into a React LoginPage",
         "figma", "frontend", "High", f"Figma: {fig}24-8"],
        ["R3", "Convert the Home (Merchant tab landing) Figma design into React",
         "figma", "frontend", "High", f"Figma: {fig}24-5"],
        ["R4", "Convert the Organisation Home / search screen Figma design into React",
         "figma", "frontend", "High", f"Figma: {fig}24-6"],
        ["R5", "Convert the Edit Basic Details screen Figma design into React",
         "figma", "frontend", "Medium", f"Figma: {fig}24-7"],
        ["R6", "Write and run unit tests for the generated backend services and controllers",
         "test", "backend", "High", "JUnit + MockMvc."],
        ["R7", "Write and run unit tests for the generated React pages",
         "test", "frontend", "Medium", "Vitest + RTL."],
        ["R8", "Review the generated backend against the coding standards",
         "review", "backend", "High", "Check layering, DTOs, security, exception handling."],
        ["R9", "Review the generated React frontend against the coding standards",
         "review", "frontend", "Medium", "Check hooks rules, list keys, API/error handling."],
    ]
    for row in sample:
        ws.append(row)
    # Widen columns a little for readability.
    for col, width in zip("ABCDEF", (6, 64, 8, 10, 9, 40)):
        ws.column_dimensions[col].width = width
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Orchestrator system prompt (LLM mode)
# ---------------------------------------------------------------------------

ORCH_SYSTEM_PROMPT = f"""\
You are the Orchestrator agent for an automated application-migration pipeline.
You coordinate three specialist subagents, which are exposed to you as tools:

- run_figma_agent : converts a Figma design into React components under {UI_REL}/.
- run_spec_agent  : converts a specification into a Spring Boot backend under {BACKEND_REL}/.
- run_test_agent  : writes and runs unit tests for the generated frontend or backend.
- run_review_agent: reviews the generated frontend or backend against the project
  coding standards and writes a review report under {REVIEWS_REL}/.
- get_subagent_status : returns the current status of every subagent.

## Your job
You are given a table of requirements (read from an Excel workbook). For each
requirement:
  1. Decide which subagent should handle it. Use the `Type` column as a strong
     hint (figma -> run_figma_agent, spec -> run_spec_agent, test -> run_test_agent,
     review -> run_review_agent). If Type is "auto", infer from the wording.
  2. Write a clear, self-contained task description for that subagent and call it.
  3. Sequence the work sensibly: generate backend (spec) and frontend (figma)
     BEFORE running their tests or reviews. Backend and frontend generation are
     independent and can be done in either order. Tests and reviews both depend on
     the generated code existing.
  4. After each subagent returns, briefly note the outcome. You may call
     get_subagent_status at any time to report progress.

## Rules
- Call exactly the subagent that fits each requirement; do not try to do the
  generation yourself - you have no filesystem tools.
- For test and review requirements, pass the correct `target` ("frontend" or "backend").
- For figma requirements, pass the figma_url if one is provided in the notes.
- When all requirements are handled, produce a final summary: which subagent
  handled which requirement, and the overall result. Then stop.

Be decisive and autonomous. Do not ask the user questions; make reasonable
decisions and proceed.
"""


# ---------------------------------------------------------------------------
# Subagent-as-tools (LLM mode)
# ---------------------------------------------------------------------------

def build_orchestrator_tools(status: StatusTracker,
                             figma_bridge: Optional[McpStdioBridge]) -> List[Tool]:
    """Wrap the three subagents (and a status query) as orchestrator tools."""

    def _figma(task: str, figma_url: str = "", **_) -> str:
        result = figma_agent.run_figma_agent(
            task, figma_url=figma_url or None, status=status, bridge=figma_bridge
        )
        return json.dumps(result)

    def _spec(task: str, spec_path: str = "", **_) -> str:
        result = spec_agent.run_spec_agent(
            task, spec_path=spec_path or None, status=status
        )
        return json.dumps(result)

    def _test(task: str, target: str = "backend", **_) -> str:
        result = test_agent.run_test_agent(task, target=target, status=status)
        return json.dumps(result)

    def _review(task: str, target: str = "backend", standards_path: str = "", **_) -> str:
        result = review_agent.run_review_agent(
            task, target=target, standards_path=standards_path or None, status=status
        )
        return json.dumps(result)

    def _status(**_) -> str:
        return json.dumps(status.snapshot(), indent=2)

    return [
        Tool(
            name="run_figma_agent",
            description=f"Run the Figma->React subagent. Generates React components under {UI_REL}/.",
            input_schema={
                "type": "object",
                "properties": {
                    "task": {"type": "string", "description": "Self-contained instruction for the figma subagent."},
                    "figma_url": {"type": "string", "description": "Optional Figma design URL with node-id."},
                },
                "required": ["task"],
            },
            handler=_figma,
        ),
        Tool(
            name="run_spec_agent",
            description=f"Run the Spec->Spring Boot subagent. Generates backend code under {BACKEND_REL}/.",
            input_schema={
                "type": "object",
                "properties": {
                    "task": {"type": "string", "description": "Self-contained instruction for the spec subagent."},
                    "spec_path": {"type": "string", "description": "Optional path to the spec file (defaults to grails-ui-demo/specs.md)."},
                },
                "required": ["task"],
            },
            handler=_spec,
        ),
        Tool(
            name="run_test_agent",
            description="Run the unit-test subagent against the generated frontend or backend.",
            input_schema={
                "type": "object",
                "properties": {
                    "task": {"type": "string"},
                    "target": {"type": "string", "enum": ["frontend", "backend"], "default": "backend"},
                },
                "required": ["task", "target"],
            },
            handler=_test,
        ),
        Tool(
            name="run_review_agent",
            description=f"Run the code-review subagent against the generated frontend or backend. "
                        f"Reviews against the coding standards and writes a report under {REVIEWS_REL}/.",
            input_schema={
                "type": "object",
                "properties": {
                    "task": {"type": "string"},
                    "target": {"type": "string", "enum": ["frontend", "backend"], "default": "backend"},
                    "standards_path": {"type": "string", "description": "Optional path to the coding standards file (defaults to mpm-agents/coding_standards.md)."},
                },
                "required": ["task", "target"],
            },
            handler=_review,
        ),
        Tool(
            name="get_subagent_status",
            description="Get the current status (state, turns, tool calls, files) of every subagent.",
            input_schema={"type": "object", "properties": {}, "required": []},
            handler=_status,
        ),
    ]


# ---------------------------------------------------------------------------
# Deterministic dispatch (no-LLM mode)
# ---------------------------------------------------------------------------

def _route(req: dict) -> str:
    """Map a requirement to a subagent name."""
    t = req["type"]
    if t in ("figma", "spec", "test", "review"):
        return {"figma": "figma_agent", "spec": "spec_agent",
                "test": "test_agent", "review": "review_agent"}[t]
    text = f"{req['requirement']} {req['notes']}".lower()
    if "figma" in text or "react" in text or "component" in text or "screen" in text:
        return "figma_agent"
    if "review" in text or "coding standard" in text or "code quality" in text:
        return "review_agent"
    if "test" in text:
        return "test_agent"
    return "spec_agent"


def dispatch_deterministic(reqs: List[dict],
                           status: StatusTracker,
                           figma_bridge: Optional[McpStdioBridge]) -> List[dict]:
    """Route each requirement to a subagent in priority/dependency order."""
    # Generation before tests/reviews; otherwise keep workbook order.
    order = {"spec_agent": 0, "figma_agent": 0, "test_agent": 1, "review_agent": 1}
    ordered = sorted(reqs, key=lambda r: order.get(_route(r), 0))

    results = []
    for req in ordered:
        agent = _route(req)
        task = req["requirement"] or f"Handle requirement {req['id']}"
        if req["notes"]:
            task += f"\nNotes: {req['notes']}"
        print(f"\n>>> [{req['id']}] -> {agent}: {req['requirement'][:80]}")
        if agent == "figma_agent":
            figma_url = _extract_figma_url(req["notes"], req["requirement"])
            res = figma_agent.run_figma_agent(task, figma_url=figma_url, status=status, bridge=figma_bridge)
        elif agent == "spec_agent":
            res = spec_agent.run_spec_agent(task, status=status)
        elif agent == "review_agent":
            target = req["target"] if req["target"] in ("frontend", "backend") else "backend"
            res = review_agent.run_review_agent(task, target=target, status=status)
        else:
            target = req["target"] if req["target"] in ("frontend", "backend") else "backend"
            res = test_agent.run_test_agent(task, target=target, status=status)
        res["requirement_id"] = req["id"]
        results.append(res)
    return results


# ---------------------------------------------------------------------------
# Orchestration entry point
# ---------------------------------------------------------------------------

def orchestrate(requirements_path: Path, use_llm: bool = True) -> dict:
    MPM_UI_DIR.mkdir(parents=True, exist_ok=True)
    MPM_BACKEND_DIR.mkdir(parents=True, exist_ok=True)

    reqs = read_requirements(requirements_path)
    if not reqs:
        print(f"No requirements found in {requirements_path}.")
        return {"ok": False, "error": "no requirements"}

    print(f"Loaded {len(reqs)} requirement(s) from {requirements_path.name}:\n")
    print(requirements_to_markdown(reqs))

    status = StatusTracker()
    # Pre-register subagents so the status table shows them from the start.
    for name in ("figma_agent", "spec_agent", "test_agent", "review_agent"):
        status.register(name)

    # One shared Figma MCP connection for all figma tasks (skipped in MOCK).
    figma_bridge: Optional[McpStdioBridge] = None
    if not MOCK and any(_route(r) == "figma_agent" for r in reqs):
        try:
            figma_bridge = load_figma_mcp_bridge()
        except Exception as e:  # noqa: BLE001
            print(f"WARNING: could not start Figma MCP server ({e}). "
                  f"figma_agent will run without it.")

    try:
        if use_llm:
            result = _orchestrate_llm(reqs, status, figma_bridge)
        else:
            results = dispatch_deterministic(reqs, status, figma_bridge)
            result = {"ok": True, "mode": "deterministic", "results": results}
    finally:
        if figma_bridge is not None:
            figma_bridge.stop()

    print("\n" + "=" * 80)
    print("FINAL SUBAGENT STATUS")
    print(status.render_table())
    print("=" * 80)
    return result


def _orchestrate_llm(reqs: List[dict],
                     status: StatusTracker,
                     figma_bridge: Optional[McpStdioBridge]) -> dict:
    tools = build_orchestrator_tools(status, figma_bridge)
    runner = AgentRunner(
        name=ORCH_NAME,
        system_prompt=ORCH_SYSTEM_PROMPT,
        tools=tools,
        status=status,
        effort="high",
    )
    task = (
        "Here are the migration requirements. Decompose them and call the "
        "appropriate subagents to fulfil each one, then summarise.\n\n"
        + requirements_to_markdown(reqs)
        + f"\n\nDefault specification file: {DEFAULT_SPEC_FILE.name} "
          f"(at grails-ui-demo/specs.md). Output: backend -> {BACKEND_REL}/, "
          f"frontend -> {UI_REL}/."
    )
    return runner.run(task)


def _cli() -> None:
    parser = argparse.ArgumentParser(description="MPM autonomous orchestrator")
    parser.add_argument("--requirements", default=str(DEFAULT_REQUIREMENTS),
                        help="Path to the requirements .xlsx workbook.")
    parser.add_argument("--no-llm", action="store_true",
                        help="Deterministic routing (no API key needed).")
    parser.add_argument("--init-sample", action="store_true",
                        help="Write a sample requirements.xlsx and exit.")
    args = parser.parse_args()

    print_banner()

    if args.init_sample:
        out = Path(args.requirements)
        write_sample_requirements(out)
        print(f"Sample requirements written to {out}")
        return

    req_path = Path(args.requirements)
    if not req_path.exists():
        print(f"Requirements file not found: {req_path}")
        print("Create one with:  python orchestrator.py --init-sample")
        return

    result = orchestrate(req_path, use_llm=not args.no_llm)
    print("\n=== orchestrator result ===")
    print(json.dumps(result, indent=2, default=str)[:4000])


if __name__ == "__main__":
    _cli()
